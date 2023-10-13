import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


## Function to create workbook from dataframe
def create_workbook(ws, df, **formatted_columns):
    """
    Populate an Excel worksheet with data from a DataFrame and apply formatting.

    This function takes an openpyxl worksheet (ws) and a pandas DataFrame (df) as input.
    It also accepts optional keyword arguments for formatting specific columns.

    Args:
        ws (openpyxl.worksheet.Worksheet): The worksheet where the data will be written.
        df (pandas.DataFrame): The DataFrame containing the data to be written to the worksheet.
        **formatted_columns: Optional keyword arguments to specify formatting for specific columns.
            Supported formats: 'date' (to format as date) and 'currency' (to format as currency).

    Returns:
        None

    Example:
        # Example usage of the create_workbook function
        create_workbook(worksheet, data_df, DateColumn="date", AmountColumn="currency")

    Note:
        - The function applies title case to column headers by default.
        - You can specify formatting for columns using the keyword arguments.
        - After applying formatting and writing data, it adjusts column widths for readability.
    """
    # Format headers
    df.columns = df.columns.str.title()

    # Format Dates within dataframe, save currency to format in worksheet
    currency_columns = []
    for col, format in formatted_columns.items():
        if format == "date":
            df[col] = df[col].dt.strftime("%m/%d/%Y")
        elif format == "currency":
            currency_columns.append(get_column_letter(df.columns.get_loc(col) + 1))

    # Put data into worksheet as atable
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    tab = Table(
        displayName=ws.title,
        ref=f"A1:{get_column_letter(ws.max_column)}{len(df)+1}",
        headerRowCount=1,
        totalsRowCount=0,
        tableStyleInfo=style,
    )

    ws.add_table(tab)

    # Format currency within worksheet
    for i in currency_columns:
        for cell in ws[i]:
            cell.number_format = '"$"#,##0.00_-'

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.4
        ws.column_dimensions[column].width = adjusted_width


## Function to format the totals rows
def format_totals(ws):
    """
    Format and highlight totals rows in an Excel worksheet.

    This function iterates through rows in an openpyxl worksheet (ws) and identifies
    totals rows based on a criteria (rows where the first cell value starts with 'E').
    It then applies specific formatting to the entire row, making it bold and changing
    the fill color for better visibility.

    Args:
        ws (openpyxl.worksheet.Worksheet): The worksheet where totals rows need to be formatted.

    Returns:
        None

    Note:
        - This function is designed to format totals rows in Excel worksheets for better clarity.
        - It identifies totals rows based on the criteria that the first cell value starts with 'E'.
        - The function doesn't return any value; it directly applies formatting to the worksheet.
    """
    # Iterate through rows and check column A values
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        cell_value = row[0].value

        # Check if the cell value contains 'Total'
        if "Total" in cell_value:
            # Apply font changes to the entire row
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", start_color="b7aea5")


if __name__ == "__main__":
    # Sample data
    data = {
        "Description": ["Salary", "Groceries", "Rent", "Total"],
        "Amount": [5000.00, -150.50, -1200.00, 3649.50],
        "Payment_Date": ["2023-10-13", "2023-10-14", "2023-10-15", None],
    }

    sample_df = pd.DataFrame(data)
    sample_df["Payment_Date"] = pd.to_datetime(
        sample_df["Payment_Date"]
    )  # Data from a database will frequently be in datetime format,
    # we'll change it to a nicely formatted string when we write to excel

    # Create workbook
    wb = Workbook()

    # Add a default table style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    ## Create Sheet
    ws = wb.active
    ws.title = "Sample_Sheet"
    create_workbook(ws, sample_df, Amount="currency", Payment_Date="date")
    format_totals(ws)

# save workbook
saving = True
while saving:
    try:
        wb.save("Sample.xlsx")
    except PermissionError:
        if input("Please close the excel file then type 'y' to continue: ") != "y":
            print("File not saved")
            saving = False
    else:
        print("File saved")
        saving = False
